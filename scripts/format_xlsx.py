# -*- coding: utf-8 -*-
"""
Apply comprehensive formatting to ShoesHub_Test_Cases.xlsx
Design goals:
  - Clear visual hierarchy with styled title/header rows
  - Color-coded Priority, Type, Automation Status
  - Wrap text preserved for multi-line cells
  - Consistent column widths and alignment
  - Frozen header rows for easy scrolling
"""

import sys
import io
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

XLSX_PATH = 'docs/ShoesHub_Test_Cases.xlsx'
wb = openpyxl.load_workbook(XLSX_PATH)

# ─── Colour Palette ───────────────────────────────────────────────────────────

C = {
    'navy':        '1A2E5A',
    'blue_dark':   '2B4590',
    'blue_mid':    '3D5FC4',
    'blue_light':  'C7D3F5',
    'blue_pale':   'EEF1FB',
    'white':       'FFFFFF',
    'off_white':   'F8F9FD',
    'gray_light':  'E9ECEF',
    'gray_mid':    'CED4DA',
    'text_dark':   '1A1A2E',
    'text_mid':    '3A3A5C',
    # priority
    'p0_bg':       'C62828',
    'p1_bg':       'E65100',
    'p2_bg':       'F9A825',
    'p3_bg':       '2E7D32',
    # type
    'type_perf':   '6A1B9A',
    'type_func':   '1565C0',
    'type_neg':    'AD1457',
    'type_smoke':  '00695C',
    'type_sec':    'B71C1C',
    'type_ui':     '0277BD',
    'type_e2e':    '00838F',
    # automation
    'auto_done':   '1B5E20',
    'auto_todo':   'E65100',
    'auto_manual': '880E4F',
    # run status
    'st_pass':     '1B5E20',
    'st_fail':     'B71C1C',
    'st_notrun':   '455A64',
    'st_flaky':    'F57F17',
    'st_skip':     '6D4C41',
}

# ─── Style helpers ────────────────────────────────────────────────────────────

def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def font(bold=False, color='1A1A2E', size=10, italic=False):
    return Font(name='Calibri', bold=bold, color=color, size=size, italic=italic)

def border_thin(sides='all'):
    thin = Side(style='thin', color=C['gray_mid'])
    none = Side(style=None)
    t = thin if 'all' in sides or 't' in sides else none
    b = thin if 'all' in sides or 'b' in sides else none
    l = thin if 'all' in sides or 'l' in sides else none
    r = thin if 'all' in sides or 'r' in sides else none
    return Border(top=t, bottom=b, left=l, right=r)

def border_outer():
    med = Side(style='medium', color=C['blue_dark'])
    return Border(top=med, bottom=med, left=med, right=med)

def align(h='left', v='top', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def apply(cell, bg=None, fg='1A1A2E', bold=False, size=10,
          h='left', v='top', wrap=True, italic=False, border=True):
    if bg:
        cell.fill = fill(bg)
    cell.font  = font(bold=bold, color=fg, size=size, italic=italic)
    cell.alignment = align(h=h, v=v, wrap=wrap)
    if border:
        cell.border = border_thin()


# ─── Sheet 1: 📋 Test Cases ───────────────────────────────────────────────────

ws = wb['📋 Test Cases']

# Column widths  (A–P)
COL_WIDTHS = {
    'A': 14,   # TC_ID
    'B': 18,   # Module
    'C': 18,   # Feature
    'D': 32,   # Test Name
    'E': 11,   # Priority
    'F': 16,   # Type
    'G': 11,   # Role
    'H': 32,   # Precondition
    'I': 48,   # Test Steps
    'J': 48,   # Expected Result
    'K': 30,   # data-testids
    'L': 32,   # API Endpoint
    'M': 14,   # Test Data Ref
    'N': 18,   # Automation Status
    'O': 32,   # Script Path
    'P': 30,   # Tags
}
for col_letter, width in COL_WIDTHS.items():
    ws.column_dimensions[col_letter].width = width

# ── Row 1: Title bar ──────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 32
ws.merge_cells('A1:P1')
title_cell = ws['A1']
title_cell.value = '📋  ShoesHub — Test Cases'
apply(title_cell, bg=C['navy'], fg=C['white'], bold=True, size=14, h='center', v='center')

# ── Row 2: Column Headers ─────────────────────────────────────────────────────
ws.row_dimensions[2].height = 28
HEADERS = [
    'TC_ID', 'Module', 'Feature', 'Test Name (TH)',
    'Priority', 'Type', 'Role', 'Precondition',
    'Test Steps', 'Expected Result', 'Key data-testids',
    'API Endpoint', 'Test Data Ref', 'Automation Status',
    'Script Path', 'Tags',
]
for col_idx, header in enumerate(HEADERS, start=1):
    cell = ws.cell(row=2, column=col_idx, value=header)
    apply(cell, bg=C['blue_dark'], fg=C['white'], bold=True, size=10,
          h='center', v='center', wrap=False)

# ── Data rows: color per module group ────────────────────────────────────────
WRAP_COLS = {'H', 'I', 'J', 'K', 'L'}  # columns that need wrap text

PRIORITY_STYLE = {
    'P0': (C['p0_bg'], C['white']),
    'P1': (C['p1_bg'], C['white']),
    'P2': (C['p2_bg'], '1A1A2E'),
    'P3': (C['p3_bg'], C['white']),
}
TYPE_STYLE = {
    'Performance': (C['type_perf'], C['white']),
    'Functional':  (C['type_func'], C['white']),
    'Negative':    (C['type_neg'],  C['white']),
    'Smoke':       (C['type_smoke'],C['white']),
    'Security':    (C['type_sec'],  C['white']),
    'UI':          (C['type_ui'],   C['white']),
    'E2E':         (C['type_e2e'],  C['white']),
}
AUTO_STYLE = {
    'Automated':        (C['auto_done'],   C['white']),
    'To Be Automated':  (C['auto_todo'],   C['white']),
    'Manual Only':      (C['auto_manual'], C['white']),
}

current_module = None
row_parity = 0  # alternates 0/1 within each module block

for row_idx in range(3, ws.max_row + 1):
    row_data = [ws.cell(row=row_idx, column=c).value for c in range(1, 17)]
    if not any(row_data):
        continue

    module = row_data[1]  # col B
    if module != current_module:
        current_module = module
        row_parity = 0
    else:
        row_parity = 1 - row_parity

    bg = C['off_white'] if row_parity else C['white']
    ws.row_dimensions[row_idx].height = 60

    for col_idx in range(1, 17):
        cell = ws.cell(row=row_idx, column=col_idx)
        col_letter = get_column_letter(col_idx)
        wrap = col_letter in WRAP_COLS
        apply(cell, bg=bg, fg=C['text_dark'], h='left', v='top', wrap=wrap)

    # Priority cell (E) — colored badge
    e_cell = ws.cell(row=row_idx, column=5)
    prio = e_cell.value
    if prio in PRIORITY_STYLE:
        bg_p, fg_p = PRIORITY_STYLE[prio]
        apply(e_cell, bg=bg_p, fg=fg_p, bold=True, h='center', v='center', wrap=False)

    # Type cell (F) — colored badge
    f_cell = ws.cell(row=row_idx, column=6)
    ttype = f_cell.value
    if ttype in TYPE_STYLE:
        bg_t, fg_t = TYPE_STYLE[ttype]
        apply(f_cell, bg=bg_t, fg=fg_t, bold=True, h='center', v='center', wrap=False)

    # Automation Status cell (N) — colored badge
    n_cell = ws.cell(row=row_idx, column=14)
    auto = n_cell.value
    if auto in AUTO_STYLE:
        bg_a, fg_a = AUTO_STYLE[auto]
        apply(n_cell, bg=bg_a, fg=fg_a, bold=True, h='center', v='center', wrap=False)

# ── Freeze pane below header ──────────────────────────────────────────────────
ws.freeze_panes = 'A3'

print('[TC] Formatted Test Cases sheet')


# ─── Sheet 2: ▶ Test Run Log ──────────────────────────────────────────────────

ws2 = wb['▶ Test Run Log']

LOG_COL_WIDTHS = {
    'A': 22, 'B': 14, 'C': 12, 'D': 12,
    'E': 12, 'F': 14, 'G': 35, 'H': 14,
    'I': 35, 'J': 14, 'K': 14, 'L': 30,
}
for col_letter, width in LOG_COL_WIDTHS.items():
    ws2.column_dimensions[col_letter].width = width

# Title
ws2.row_dimensions[1].height = 32
ws2.merge_cells('A1:L1')
t = ws2['A1']
t.value = '▶  ShoesHub — Test Run Log'
apply(t, bg=C['navy'], fg=C['white'], bold=True, size=14, h='center', v='center')

# Headers
ws2.row_dimensions[2].height = 28
LOG_HEADERS = [
    'Run ID', 'Run Date', 'Environment', 'Branch', 'Tester',
    'TC_ID', 'Test Name', 'Status', 'Actual Result', 'Bug ID',
    'Duration (ms)', 'Notes',
]
for col_idx, header in enumerate(LOG_HEADERS, start=1):
    cell = ws2.cell(row=2, column=col_idx, value=header)
    apply(cell, bg=C['blue_dark'], fg=C['white'], bold=True,
          h='center', v='center', wrap=False)

STATUS_STYLE = {
    'Pass':    (C['st_pass'],   C['white']),
    'Fail':    (C['st_fail'],   C['white']),
    'Not Run': (C['st_notrun'], C['white']),
    'Flaky':   (C['st_flaky'],  C['white']),
    'Skipped': (C['st_skip'],   C['white']),
}

for row_idx in range(3, ws2.max_row + 1):
    row_data = [ws2.cell(row=row_idx, column=c).value for c in range(1, 13)]
    if not any(row_data):
        continue

    bg = C['off_white'] if row_idx % 2 == 0 else C['white']
    ws2.row_dimensions[row_idx].height = 22

    for col_idx in range(1, 13):
        cell = ws2.cell(row=row_idx, column=col_idx)
        wrap = col_idx in (7, 9, 12)  # Test Name, Actual Result, Notes
        apply(cell, bg=bg, fg=C['text_dark'], h='left', v='center', wrap=wrap)

    # Status cell (H, col 8)
    status_cell = ws2.cell(row=row_idx, column=8)
    status = status_cell.value
    if status in STATUS_STYLE:
        bg_s, fg_s = STATUS_STYLE[status]
        apply(status_cell, bg=bg_s, fg=fg_s, bold=True,
              h='center', v='center', wrap=False)

ws2.freeze_panes = 'A3'
print('[Log] Formatted Test Run Log sheet')


# ─── Sheet 3: 📁 Test Data ────────────────────────────────────────────────────

ws3 = wb['📁 Test Data']
TD_WIDTHS = {'A': 16, 'B': 16, 'C': 20, 'D': 16, 'E': 24, 'F': 16, 'G': 16, 'H': 20, 'I': 25}
for col, w in TD_WIDTHS.items():
    ws3.column_dimensions[col].width = w

ws3.row_dimensions[1].height = 32
ws3.merge_cells('A1:I1')
t = ws3['A1']
t.value = '📁  ShoesHub — Test Data'
apply(t, bg=C['navy'], fg=C['white'], bold=True, size=14, h='center', v='center')

ws3.row_dimensions[2].height = 28
for cell in ws3[2]:
    apply(cell, bg=C['blue_dark'], fg=C['white'], bold=True,
          h='center', v='center', wrap=False)

for row_idx in range(3, ws3.max_row + 1):
    row_data = [ws3.cell(row=row_idx, column=c).value for c in range(1, 10)]
    if not any(row_data):
        continue
    bg = C['off_white'] if row_idx % 2 == 0 else C['white']
    ws3.row_dimensions[row_idx].height = 20
    for col_idx in range(1, 10):
        cell = ws3.cell(row=row_idx, column=col_idx)
        apply(cell, bg=bg, fg=C['text_dark'], v='center')

ws3.freeze_panes = 'A3'
print('[Data] Formatted Test Data sheet')


# ─── Sheet 4: 🌐 Environments ─────────────────────────────────────────────────

ws4 = wb['🌐 Environments']
ENV_WIDTHS = {'A': 14, 'B': 30, 'C': 30, 'D': 14, 'E': 14,
              'F': 14, 'G': 14, 'H': 10, 'I': 12, 'J': 35}
for col, w in ENV_WIDTHS.items():
    ws4.column_dimensions[col].width = w

ws4.row_dimensions[1].height = 32
ws4.merge_cells('A1:J1')
t = ws4['A1']
t.value = '🌐  Environments'
apply(t, bg=C['navy'], fg=C['white'], bold=True, size=14, h='center', v='center')

ws4.row_dimensions[2].height = 28
for cell in ws4[2]:
    apply(cell, bg=C['blue_dark'], fg=C['white'], bold=True,
          h='center', v='center', wrap=False)

ENV_ROW_COLORS = ['2D6A4F', '1565C0', '6A1B9A', 'AD1457']  # dev, qa, stg, prod
env_color_idx = 0
for row_idx in range(3, ws4.max_row + 1):
    row_data = [ws4.cell(row=row_idx, column=c).value for c in range(1, 11)]
    if not any(row_data):
        continue
    ws4.row_dimensions[row_idx].height = 22
    env_name = str(row_data[0] or '').upper()
    row_bg = C['off_white'] if row_idx % 2 == 0 else C['white']
    for col_idx in range(1, 11):
        cell = ws4.cell(row=row_idx, column=col_idx)
        apply(cell, bg=row_bg, fg=C['text_dark'], v='center')
    # Highlight env name in col A
    a_cell = ws4.cell(row=row_idx, column=1)
    if env_color_idx < len(ENV_ROW_COLORS):
        apply(a_cell, bg=ENV_ROW_COLORS[env_color_idx], fg=C['white'],
              bold=True, h='center', v='center', wrap=False)
    env_color_idx += 1

ws4.freeze_panes = 'A3'
print('[Env] Formatted Environments sheet')


# ─── Sheet 5: 🐛 Bug Report ───────────────────────────────────────────────────

ws5 = wb['🐛 Bug Report']
BUG_WIDTHS = {
    'A': 14, 'B': 12, 'C': 14, 'D': 12, 'E': 12, 'F': 10,
    'G': 35, 'H': 40, 'I': 30, 'J': 30, 'K': 14, 'L': 14,
    'M': 14, 'N': 30,
}
for col, w in BUG_WIDTHS.items():
    ws5.column_dimensions[col].width = w

ws5.row_dimensions[1].height = 32
ws5.merge_cells('A1:N1')
t = ws5['A1']
t.value = '🐛  Bug Report'
apply(t, bg=C['p0_bg'], fg=C['white'], bold=True, size=14, h='center', v='center')

ws5.row_dimensions[2].height = 28
for cell in ws5[2]:
    apply(cell, bg='7B1818', fg=C['white'], bold=True,
          h='center', v='center', wrap=False)

SEV_STYLE = {
    'Critical': ('B71C1C', C['white']),
    'High':     ('E65100', C['white']),
    'Medium':   ('F9A825', '1A1A2E'),
    'Low':      ('2E7D32', C['white']),
}
for row_idx in range(3, ws5.max_row + 1):
    row_data = [ws5.cell(row=row_idx, column=c).value for c in range(1, 15)]
    if not any(row_data):
        continue
    bg = C['off_white'] if row_idx % 2 == 0 else C['white']
    ws5.row_dimensions[row_idx].height = 20
    for col_idx in range(1, 15):
        cell = ws5.cell(row=row_idx, column=col_idx)
        wrap = col_idx in (7, 8, 9, 10, 14)
        apply(cell, bg=bg, fg=C['text_dark'], v='center', wrap=wrap)
    # Severity (E, col 5)
    sev_cell = ws5.cell(row=row_idx, column=5)
    sev = sev_cell.value
    if sev in SEV_STYLE:
        bg_s, fg_s = SEV_STYLE[sev]
        apply(sev_cell, bg=bg_s, fg=fg_s, bold=True, h='center', v='center', wrap=False)

ws5.freeze_panes = 'A3'
print('[Bug] Formatted Bug Report sheet')


# ─── Sheet 6: 📊 Summary ─────────────────────────────────────────────────────

ws6 = wb['📊 Summary']

# Clear existing content and rebuild
ws6.column_dimensions['A'].width = 22
ws6.column_dimensions['B'].width = 10
ws6.column_dimensions['C'].width = 4
ws6.column_dimensions['D'].width = 22
ws6.column_dimensions['E'].width = 10

# Rebuild summary data
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def sum_title(cell, text, bg, fg=C['white']):
    cell.value = text
    cell.fill = fill(bg)
    cell.font = font(bold=True, color=fg, size=11)
    cell.alignment = align(h='center', v='center', wrap=False)
    cell.border = border_thin()

def sum_label(cell, text, bg=C['blue_pale'], fg=C['text_dark'], bold=False):
    cell.value = text
    cell.fill = fill(bg)
    cell.font = font(bold=bold, color=fg, size=10)
    cell.alignment = align(h='left', v='center', wrap=False)
    cell.border = border_thin()

def sum_count(cell, value, bg=C['white'], fg=C['text_dark'], bold=False):
    cell.value = value
    cell.fill = fill(bg)
    cell.font = font(bold=bold, color=fg, size=10)
    cell.alignment = align(h='center', v='center', wrap=False)
    cell.border = border_thin()

# Clear the sheet first — unmerge all, then reset cells
from openpyxl.cell.cell import MergedCell
for rng in list(ws6.merged_cells.ranges):
    ws6.unmerge_cells(str(rng))
for row in ws6.iter_rows():
    for cell in row:
        if isinstance(cell, MergedCell):
            continue
        cell.value = None
        cell.fill = PatternFill()
        cell.font = Font()
        cell.border = Border()
        cell.alignment = Alignment()

# Build layout
# Row 1: Title
ws6.row_dimensions[1].height = 34
ws6.merge_cells('A1:E1')
t = ws6['A1']
t.value = '📊  Test Summary — ShoesHub'
apply(t, bg=C['navy'], fg=C['white'], bold=True, size=14, h='center', v='center')

# Row 2: Total
ws6.row_dimensions[2].height = 26
ws6.merge_cells('A2:E2')
tc = ws6['A2']
tc.value = 'Total Test Cases: 92'
apply(tc, bg=C['blue_mid'], fg=C['white'], bold=True, size=12, h='center', v='center')

# Row 3: section headers
ws6.row_dimensions[3].height = 22
sum_title(ws6['A3'], 'By Module', C['blue_dark'])
sum_title(ws6['D3'], 'By Priority', C['blue_dark'])

# Module rows
modules = [
    ('Admin Dashboard', 3), ('Admin Orders', 6), ('Admin Products', 7),
    ('Authentication', 10), ('Cart', 6), ('Checkout', 5),
    ('Homepage', 3), ('Navigation', 7), ('Orders', 4),
    ('Performance', 14), ('Product Detail', 6), ('Products', 9),
    ('Profile', 7), ('i18n', 5),
]
MODULE_COLORS = {
    'Authentication': '1565C0', 'Products': '00695C', 'Product Detail': '00695C',
    'Cart': 'AD1457', 'Checkout': 'AD1457', 'Orders': '4A148C',
    'Profile': '1B5E20', 'Admin Dashboard': '4E342E', 'Admin Products': '4E342E',
    'Admin Orders': '4E342E', 'Navigation': '37474F', 'Homepage': '37474F',
    'i18n': '37474F', 'Performance': '6A1B9A',
}

priority_rows = [
    ('P0 — Smoke',   12, C['p0_bg'],  C['white']),
    ('P1 — High',    44, C['p1_bg'],  C['white']),
    ('P2 — Medium',  32, C['p2_bg'],  '1A1A2E'),
    ('P3 — Low',      4, C['p3_bg'],  C['white']),
]

for i, (mod, cnt) in enumerate(modules):
    r = 4 + i
    ws6.row_dimensions[r].height = 20
    col = MODULE_COLORS.get(mod, C['blue_mid'])
    sum_label(ws6.cell(r, 1), mod, bg=col, fg=C['white'], bold=True)
    sum_count(ws6.cell(r, 2), cnt, bg=C['off_white'], bold=True)

    if i < len(priority_rows):
        lbl, val, pbg, pfg = priority_rows[i]
        sum_label(ws6.cell(r, 4), lbl, bg=pbg, fg=pfg, bold=True)
        sum_count(ws6.cell(r, 5), val, bg=C['off_white'], bold=True)

# Blank separator row
sep_r = 4 + len(modules)
ws6.row_dimensions[sep_r].height = 10

# By Type / Automation Status headers
t_r = sep_r + 1
ws6.row_dimensions[t_r].height = 22
sum_title(ws6['A' + str(t_r)], 'By Type', C['blue_dark'])
sum_title(ws6['D' + str(t_r)], 'Automation Status', C['blue_dark'])

types = [
    ('Functional',   44, C['type_func'],  C['white']),
    ('Smoke',         9, C['type_smoke'], C['white']),
    ('Performance',  14, C['type_perf'],  C['white']),
    ('UI',           11, C['type_ui'],    C['white']),
    ('Negative',     10, C['type_neg'],   C['white']),
    ('Security',      3, C['type_sec'],   C['white']),
    ('E2E',           1, C['type_e2e'],   C['white']),
]
auto_rows = [
    ('Automated',       76, C['auto_done'],   C['white']),
    ('To Be Automated',  1, C['auto_todo'],   C['white']),
    ('Manual Only',      1, C['auto_manual'], C['white']),
]

for i, (ttype, cnt, tbg, tfg) in enumerate(types):
    r = t_r + 1 + i
    ws6.row_dimensions[r].height = 20
    sum_label(ws6.cell(r, 1), ttype, bg=tbg, fg=tfg, bold=True)
    sum_count(ws6.cell(r, 2), cnt, bg=C['off_white'], bold=True)

    if i < len(auto_rows):
        albl, aval, abg, afg = auto_rows[i]
        sum_label(ws6.cell(r, 4), albl, bg=abg, fg=afg, bold=True)
        sum_count(ws6.cell(r, 5), aval, bg=C['off_white'], bold=True)

print('[Summary] Rebuilt Summary sheet')


# ─── Sheet 7: 📖 Legend ──────────────────────────────────────────────────────

ws7 = wb['📖 Legend']
ws7.column_dimensions['A'].width = 22
ws7.column_dimensions['B'].width = 40
ws7.column_dimensions['C'].width = 4
ws7.column_dimensions['D'].width = 22
ws7.column_dimensions['E'].width = 40

for rng in list(ws7.merged_cells.ranges):
    ws7.unmerge_cells(str(rng))
for row in ws7.iter_rows():
    for cell in row:
        if isinstance(cell, MergedCell):
            continue
        cell.value = None
        cell.fill = PatternFill()
        cell.font = Font()
        cell.border = Border()
        cell.alignment = Alignment()

ws7.row_dimensions[1].height = 34
ws7.merge_cells('A1:E1')
t = ws7['A1']
t.value = '📖  Legend & Color Guide'
apply(t, bg=C['navy'], fg=C['white'], bold=True, size=14, h='center', v='center')

def leg_section(row, col_a, text):
    ws7.row_dimensions[row].height = 22
    c = ws7.cell(row=row, column=col_a)
    ws7.merge_cells(
        start_row=row, start_column=col_a,
        end_row=row,   end_column=col_a + 1
    )
    apply(c, bg=C['blue_dark'], fg=C['white'], bold=True, h='center', v='center')
    c.value = text

def leg_row(row, col_a, label, desc, bg_label, fg_label=C['white']):
    ws7.row_dimensions[row].height = 20
    lc = ws7.cell(row=row, column=col_a)
    apply(lc, bg=bg_label, fg=fg_label, bold=True, h='center', v='center', wrap=False)
    lc.value = label
    dc = ws7.cell(row=row, column=col_a + 1)
    apply(dc, bg=C['off_white'], fg=C['text_dark'], h='left', v='center', wrap=False)
    dc.value = desc

r = 2
leg_section(r, 1, 'PRIORITY')
leg_section(r, 4, 'AUTOMATION STATUS')
r += 1
leg_row(r, 1, 'P0', 'Smoke — ระบบพังถ้าไม่ผ่าน', C['p0_bg'])
leg_row(r, 4, 'Automated', 'มี script รันอัตโนมัติแล้ว', C['auto_done'])
r += 1
leg_row(r, 1, 'P1', 'High — ฟีเจอร์หลักที่สำคัญ', C['p1_bg'])
leg_row(r, 4, 'To Be Automated', 'วางแผน automate แต่ยังไม่ทำ', C['auto_todo'])
r += 1
leg_row(r, 1, 'P2', 'Medium — ฟีเจอร์รอง', C['p2_bg'], fg_label='1A1A2E')
leg_row(r, 4, 'Manual Only', 'ทดสอบมือเท่านั้น', C['auto_manual'])
r += 1
leg_row(r, 1, 'P3', 'Low — Nice-to-have', C['p3_bg'])
r += 1
r += 1  # spacer

leg_section(r, 1, 'TEST TYPE')
leg_section(r, 4, 'RUN STATUS')
r += 1
leg_row(r, 1, 'Functional', 'ทดสอบ business logic', C['type_func'])
leg_row(r, 4, 'Pass', 'ผ่าน', C['st_pass'])
r += 1
leg_row(r, 1, 'Smoke', 'Sanity check จำนวนน้อย', C['type_smoke'])
leg_row(r, 4, 'Fail', 'ไม่ผ่าน มี bug', C['st_fail'])
r += 1
leg_row(r, 1, 'Performance', 'ทดสอบ throughput/latency', C['type_perf'])
leg_row(r, 4, 'Not Run', 'ยังไม่ได้รัน', C['st_notrun'])
r += 1
leg_row(r, 1, 'Negative', 'ทดสอบ invalid input', C['type_neg'])
leg_row(r, 4, 'Flaky', 'ผลไม่สม่ำเสมอ ต้องตรวจสอบ', C['st_flaky'])
r += 1
leg_row(r, 1, 'Security', 'ทดสอบ auth/authz/injection', C['type_sec'])
leg_row(r, 4, 'Skipped', 'ข้ามโดยตั้งใจ', C['st_skip'])
r += 1
leg_row(r, 1, 'UI', 'ทดสอบ layout/visual', C['type_ui'])
r += 1
leg_row(r, 1, 'E2E', 'End-to-end user journey', C['type_e2e'])

print('[Legend] Rebuilt Legend sheet')


# ─── Save ─────────────────────────────────────────────────────────────────────

wb.save(XLSX_PATH)
print(f'\nDone. Saved to {XLSX_PATH}')
