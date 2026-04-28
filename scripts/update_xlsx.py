# -*- coding: utf-8 -*-
"""
Safe data-update utilities for ShoesHub_Test_Cases.xlsx.

RULE: Existing rows — modify ONLY .value, never .fill/.font/.alignment/.border.
      openpyxl preserves all formatting as long as style properties are untouched.
      Line breaks (\n) in cell values are also preserved automatically.

      New rows — call `python scripts/format_xlsx.py` after adding them so the
      formatter can apply the correct styles to the new content.

Usage examples (run from repo root):
  python scripts/update_xlsx.py status TC-AUTH-01 Automated tests/auth/TC_AUTH_001.spec.ts
  python scripts/update_xlsx.py log 2024-01-15 staging main tester@mail.com
"""

import sys
import io
import openpyxl
from datetime import date

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

XLSX_PATH = 'docs/ShoesHub_Test_Cases.xlsx'

VALID_AUTO_STATUS = {'Automated', 'To Be Automated', 'Manual Only'}
VALID_RUN_STATUS  = {'Pass', 'Fail', 'Not Run', 'Flaky', 'Skipped'}


def _load():
    return openpyxl.load_workbook(XLSX_PATH)


def _save(wb):
    wb.save(XLSX_PATH)
    print(f'Saved → {XLSX_PATH}')
    print('Tip: run `python scripts/format_xlsx.py` if you added new rows.')


# ─── 1. Update Automation Status + Script Path ────────────────────────────────

def update_automation_status(tc_id: str, status: str, script_path: str = ''):
    """Update columns N (Automation Status) and O (Script Path) for a TC.

    Only modifies .value — formatting is untouched.
    """
    if status not in VALID_AUTO_STATUS:
        raise ValueError(f'Invalid status "{status}". Use: {VALID_AUTO_STATUS}')

    wb = _load()
    ws = wb['📋 Test Cases']
    found = False
    for row in ws.iter_rows(min_row=3):
        if row[0].value == tc_id:
            row[13].value = status       # col N — Automation Status
            row[14].value = script_path  # col O — Script Path
            found = True
            break
    if not found:
        raise ValueError(f'TC_ID "{tc_id}" not found in Test Cases sheet')

    _save(wb)
    print(f'[status] {tc_id} → {status}')


# ─── 2. Bulk update Automation Status from a dict ─────────────────────────────

def bulk_update_automation(updates: dict):
    """Update multiple TCs at once.

    updates = {
        'TC-AUTH-01': ('Automated', 'tests/auth/TC_AUTH_001.spec.ts'),
        'TC-PROD-01': ('Automated', 'tests/products/TC_PROD_001.spec.ts'),
    }
    """
    wb = _load()
    ws = wb['📋 Test Cases']
    updated = []
    for row in ws.iter_rows(min_row=3):
        tc_id = row[0].value
        if tc_id in updates:
            status, path = updates[tc_id]
            row[13].value = status
            row[14].value = path
            updated.append(tc_id)

    missing = set(updates) - set(updated)
    if missing:
        print(f'[warn] TC_IDs not found: {missing}')

    _save(wb)
    print(f'[bulk] Updated {len(updated)} rows: {updated}')


# ─── 3. Add a new Test Run Log entry ─────────────────────────────────────────

def add_run_log_entry(
    run_id: str,
    tc_id: str,
    test_name: str,
    status: str,
    run_date: str = '',
    environment: str = 'Staging',
    branch: str = 'main',
    tester: str = '',
    actual_result: str = '',
    bug_id: str = '',
    duration_ms: str = '',
    notes: str = '',
):
    """Append a row to Test Run Log.

    New rows won't have color formatting until you run format_xlsx.py.
    """
    if status not in VALID_RUN_STATUS:
        raise ValueError(f'Invalid status "{status}". Use: {VALID_RUN_STATUS}')

    run_date = run_date or str(date.today())
    wb = _load()
    ws = wb['▶ Test Run Log']

    next_row = ws.max_row + 1
    row_data = [
        run_id, run_date, environment, branch, tester,
        tc_id, test_name, status, actual_result, bug_id,
        duration_ms, notes,
    ]
    for col_idx, val in enumerate(row_data, start=1):
        ws.cell(row=next_row, column=col_idx).value = val

    _save(wb)
    print(f'[log] Added row {next_row}: {tc_id} — {status}')


# ─── 4. Update existing run log status ────────────────────────────────────────

def update_run_log_status(run_id: str, tc_id: str, status: str,
                          actual_result: str = '', duration_ms: str = ''):
    """Update the Status (col H) of an existing run-log entry.

    Matched by both Run ID (col A) and TC_ID (col F).
    Only modifies .value — formatting is untouched.
    """
    if status not in VALID_RUN_STATUS:
        raise ValueError(f'Invalid status "{status}". Use: {VALID_RUN_STATUS}')

    wb = _load()
    ws = wb['▶ Test Run Log']
    found = False
    for row in ws.iter_rows(min_row=3):
        if row[0].value == run_id and row[5].value == tc_id:
            row[7].value  = status          # col H — Status
            if actual_result:
                row[8].value = actual_result    # col I — Actual Result
            if duration_ms:
                row[10].value = duration_ms     # col K — Duration
            found = True
            break
    if not found:
        raise ValueError(f'Entry run_id="{run_id}" tc_id="{tc_id}" not found')

    _save(wb)
    print(f'[log] Updated {run_id}/{tc_id} → {status}')


# ─── 5. Safe value-only cell update (generic) ─────────────────────────────────

def set_cell_value(sheet_name: str, row: int, col: int, value):
    """Set a single cell value without touching formatting.

    Row/col are 1-based (same as openpyxl convention).
    """
    wb = _load()
    ws = wb[sheet_name]
    ws.cell(row=row, column=col).value = value
    _save(wb)
    print(f'[cell] {sheet_name}!{row},{col} = {value!r}')


# ─── CLI entry point ──────────────────────────────────────────────────────────

if __name__ == '__main__':
    args = sys.argv[1:]
    if not args:
        print(__doc__)
        sys.exit(0)

    cmd = args[0]
    if cmd == 'status' and len(args) >= 3:
        # update_automation_status <tc_id> <status> [script_path]
        update_automation_status(args[1], args[2], args[3] if len(args) > 3 else '')
    elif cmd == 'log' and len(args) >= 4:
        # log <run_id> <tc_id> <status> [actual_result]
        add_run_log_entry(
            run_id=args[1], tc_id=args[2], status=args[3],
            actual_result=args[4] if len(args) > 4 else '',
        )
    else:
        print(__doc__)
        sys.exit(1)
